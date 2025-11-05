import os

from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def create_directory(reporting_month, directory_name, file_name):
    # Get the user's desktop directory
    desktop = os.path.join(os.path.expanduser("~"), "Desktop")

    # Specify the folder on the desktop
    folder_name = f"{directory_name}_{reporting_month.replace('/', '-')}"
    output_folder = os.path.join(desktop, folder_name)
    # Create the folder if it doesn't exist
    os.makedirs(output_folder, exist_ok=True)
    # Build the filename dynamically
    # file_name = f"{file_name}_{reporting_month.replace('/', '-')}.xlsx"
    file_name = f"{file_name}.xlsx"
    return os.path.join(output_folder, file_name)


def update_excel_file(output_file, sheets_to_ignore, sheets_to_update):
    try:
        workbook = load_workbook(output_file)
    except FileNotFoundError:
        # If file doesn't exist, create a new workbook
        workbook = Workbook()
        # Remove the default empty sheet if it exists
        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])
    #    Skip the formula-based sheets entirely
    for sheet_name, df in sheets_to_update.items():
        # If this sheet is also one of the formula sheets,
        # we skip it to avoid overwriting formulas
        if sheet_name in sheets_to_ignore:
            continue  # do not touch it

        # If the sheet already exists, clear it. Otherwise, create it.
        if sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            ws.delete_rows(1, ws.max_row)
        else:
            ws = workbook.create_sheet(title=sheet_name)

        # Write the DataFrame into the sheet (headers + data)
        for row_index, row in enumerate(
            dataframe_to_rows(df, index=False, header=True),
            start=1
        ):
            for col_index, value in enumerate(row, start=1):
                ws.cell(row=row_index, column=col_index, value=value)

    # 7. Finally, save the workbook
    workbook.save(output_file)
